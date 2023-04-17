from datetime import date
from openpyxl import load_workbook

from models import Company, FactQliq, FactQoil, ForecastQliq, ForecastQoil, db

FILE_PATH = 'task.xlsx'


def create_db():
    try:
        db.connect()
        db.create_tables(
            [Company, FactQliq, FactQoil, ForecastQliq, ForecastQoil],
        )
    except Exception as err:
        print('ERROR: сreating db, cause: {0}'.format(err))
        exit()


def load_data():
    try:
        wb = load_workbook(FILE_PATH)
    except Exception as err:
        print('ERROR: opening file, cause: {0}'.format(err))
        exit()

    ws = wb.active

    cell_range = ws['A4':'J23']

    for row in cell_range:
        for cell in row:
            if cell.column_letter == 'A':
                day_num = cell.value
            elif cell.column_letter == 'B':
                company, _ = Company.get_or_create(name=cell.value)
            elif cell.column_letter == 'C':
                data1 = cell.value
            elif cell.column_letter == 'D':
                FactQliq.create(
                    company_id=company.id,
                    data1=data1,
                    data2=cell.value,
                    total=data1 + cell.value,
                    date=date(1960, 1, day_num),
                )
            elif cell.column_letter == 'E':
                data1 = cell.value
            elif cell.column_letter == 'F':
                FactQoil.create(
                    company_id=company.id,
                    data1=data1,
                    data2=cell.value,
                    total=data1 + cell.value,
                    date=date(1960, 1, day_num),
                )
            elif cell.column_letter == 'G':
                data1 = cell.value
            elif cell.column_letter == 'H':
                ForecastQliq.create(
                    company_id=company.id,
                    data1=data1,
                    data2=cell.value,
                    total=data1 + cell.value,
                    date=date(1960, 1, day_num),
                )
            elif cell.column_letter == 'I':
                data1 = cell.value
            elif cell.column_letter == 'J':
                ForecastQoil.create(
                    company_id=company.id,
                    data1=data1,
                    data2=cell.value,
                    total=data1 + cell.value,
                    date=date(1960, 1, day_num),
                )


def main():
    create_db()
    load_data()


if __name__ == '__main__':
    main()
